"""Tests for file_parser.py — CSV and Excel parsing."""

import io
from datetime import datetime

import pytest

from src.engine.file_parser import (
    ParseError,
    ParseResult,
    parse_csv,
    parse_excel,
    parse_file,
)


class TestParseCSV:
    def test_parses_csv_with_correct_headers_and_rows(self):
        """Happy path: CSV with known headers parses to correct shape."""
        csv_bytes = (
            b"product_name,brand,upc\n"
            b"Hot Sauce,Cinderhaven,049000004502\n"
            b"Mustard,Cinderhaven,049000004519\n"
        )
        result = parse_csv(csv_bytes)

        assert isinstance(result, ParseResult)
        assert result.format == "csv"
        assert result.headers == ["product_name", "brand", "upc"]
        assert result.row_count == 2
        assert len(result.rows) == 2
        assert result.rows[0]["product_name"] == "Hot Sauce"
        assert result.rows[1]["upc"] == "049000004519"

    def test_raises_when_file_is_empty(self):
        """Edge case: empty bytes raises ParseError, not a crash."""
        with pytest.raises(ParseError, match="empty"):
            parse_csv(b"")

    def test_raises_when_file_is_whitespace_only(self):
        """Edge case: whitespace-only file treated as empty."""
        with pytest.raises(ParseError, match="empty"):
            parse_csv(b"   \n  \n  ")

    def test_returns_headers_with_empty_rows_when_headers_only(self):
        """Edge case: headers present but no data rows."""
        csv_bytes = b"product_name,brand,upc\n"
        result = parse_csv(csv_bytes)

        assert result.headers == ["product_name", "brand", "upc"]
        assert result.rows == []
        assert result.row_count == 0

    def test_handles_bom_encoding(self):
        """Edge case: CSV with UTF-8 BOM (from Excel export) parses correctly."""
        # UTF-8 BOM: \xef\xbb\xbf
        csv_bytes = b"\xef\xbb\xbfproduct_name,brand\nHot Sauce,Cinderhaven\n"
        result = parse_csv(csv_bytes)

        # BOM should be stripped — first header should be clean
        assert result.headers[0] == "product_name"
        assert result.rows[0]["product_name"] == "Hot Sauce"

    def test_handles_latin1_encoding(self):
        """Edge case: non-UTF8 file falls back to latin-1 decoding."""
        # \xe9 is 'e-acute' in latin-1, invalid as standalone UTF-8
        csv_bytes = b"name,brand\nCaf\xe9,TestBrand\n"
        result = parse_csv(csv_bytes)

        assert result.headers == ["name", "brand"]
        assert result.row_count == 1

    def test_empty_cell_values_become_none(self):
        """Empty cells in data rows become None, not empty string."""
        csv_bytes = b"name,brand,upc\nHot Sauce,,049000004502\n"
        result = parse_csv(csv_bytes)

        assert result.rows[0]["brand"] is None

    def test_strips_whitespace_from_values(self):
        """Leading/trailing whitespace in values is stripped."""
        csv_bytes = b"name,brand\n  Hot Sauce  ,  Cinderhaven  \n"
        result = parse_csv(csv_bytes)

        assert result.rows[0]["name"] == "Hot Sauce"
        assert result.rows[0]["brand"] == "Cinderhaven"


class TestParseExcel:
    def _make_xlsx(
        self,
        headers: list[str],
        rows: list[list],
    ) -> bytes:
        """Create minimal .xlsx bytes for testing."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_parses_excel_with_correct_headers_and_rows(self):
        """Happy path: Excel file with known data parses correctly."""
        xlsx = self._make_xlsx(
            ["product_name", "brand", "upc"],
            [
                ["Hot Sauce", "Cinderhaven", "049000004502"],
                ["Mustard", "Cinderhaven", "049000004519"],
            ],
        )
        result = parse_excel(xlsx)

        assert isinstance(result, ParseResult)
        assert result.format == "xlsx"
        assert result.headers == ["product_name", "brand", "upc"]
        assert result.row_count == 2
        assert result.rows[0]["product_name"] == "Hot Sauce"
        assert result.rows[1]["upc"] == "049000004519"

    def test_raises_when_file_is_empty(self):
        """Edge case: empty bytes raises ParseError."""
        with pytest.raises(ParseError, match="empty"):
            parse_excel(b"")

    def test_returns_headers_with_empty_rows_when_headers_only(self):
        """Edge case: Excel with headers row but no data."""
        xlsx = self._make_xlsx(["product_name", "brand"], [])
        result = parse_excel(xlsx)

        assert result.headers == ["product_name", "brand"]
        assert result.rows == []
        assert result.row_count == 0

    def test_converts_date_cells_to_iso_strings(self):
        """Edge case: datetime cells become ISO-format strings."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "launch_date"])
        ws.append(["Hot Sauce", datetime(2026, 3, 15)])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_bytes = buf.read()

        result = parse_excel(xlsx_bytes)

        date_val = result.rows[0]["launch_date"]
        assert date_val is not None
        assert "2026-03-15" in date_val

    def test_none_cells_become_none(self):
        """None/empty cells in Excel become None in the result."""
        xlsx = self._make_xlsx(
            ["name", "brand", "upc"],
            [["Hot Sauce", None, "049000004502"]],
        )
        result = parse_excel(xlsx)

        assert result.rows[0]["brand"] is None

    def test_raises_when_invalid_xlsx(self):
        """Non-xlsx bytes produce ParseError, not an unhandled crash."""
        with pytest.raises(ParseError, match="Cannot read"):
            parse_excel(b"this is not an xlsx file")


class TestParseFile:
    def test_dispatches_csv_by_extension(self):
        """parse_file routes .csv to parse_csv."""
        csv_bytes = b"name,brand\nHot Sauce,Cinderhaven\n"
        result = parse_file(csv_bytes, "products.csv")

        assert result.format == "csv"
        assert result.headers == ["name", "brand"]

    def test_dispatches_xlsx_by_extension(self):
        """parse_file routes .xlsx to parse_excel."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "brand"])
        ws.append(["Hot Sauce", "Cinderhaven"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        result = parse_file(buf.read(), "products.xlsx")

        assert result.format == "xlsx"
        assert result.headers == ["name", "brand"]

    def test_raises_for_unsupported_extension(self):
        """Unsupported extension raises ParseError with clear message."""
        with pytest.raises(ParseError, match="Unsupported"):
            parse_file(b"some data", "products.json")

    def test_case_insensitive_extension(self):
        """Extension detection is case-insensitive."""
        csv_bytes = b"name,brand\nHot Sauce,Cinderhaven\n"
        result = parse_file(csv_bytes, "PRODUCTS.CSV")

        assert result.format == "csv"
